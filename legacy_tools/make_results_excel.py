#!/usr/bin/env python3
"""
実験結果をExcelファイルとして出力する
"""
import json, csv, datetime
from pathlib import Path
from collections import Counter

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    subprocess.run(['/tmp/sotsuron_env312/bin/pip', 'install', 'openpyxl', '-q'])
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

WORKDIR = Path('/Users/h-torii4649/Downloads/sotsuron_latex_set')

def load_csv(path):
    records = []
    with open(path) as f:
        reader = csv.DictReader(f)
        for row in reader:
            year = int(row['year']) if row['year'] else None
            records.append({
                'year': year,
                'family_idx': row['family_idx'],
                'is_unet': row['is_unet'] == '1'
            })
    return records

def header_style(cell, bold=True, bg='2196F3', font_color='FFFFFF'):
    cell.font = Font(bold=bold, color=font_color, name='Calibri')
    cell.fill = PatternFill('solid', start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center')

def data_style(cell, number_format=None):
    cell.font = Font(name='Calibri')
    cell.alignment = Alignment(horizontal='center')
    if number_format:
        cell.number_format = number_format

def make_excel(a0_records, b0_records, summary, out_path):
    wb = Workbook()

    # ===== Sheet 1: 全体集計 =====
    ws1 = wb.active
    ws1.title = '全体集計'

    ws1['A1'] = '特許データ実験結果サマリー'
    ws1['A1'].font = Font(bold=True, size=14, name='Calibri')
    ws1['A1'] = '特許データ実験結果サマリー（U-Net クロスドメイン分析）'

    ws1.append([])
    ws1.append(['分析日', datetime.date.today().isoformat()])
    ws1.append(['対象期間', '2015年〜2024年（Filing date基準）'])
    ws1.append(['参照分野', '医療画像解析 (A0)'])
    ws1.append(['対象分野', '製造欠陥検出 (B0)'])
    ws1.append(['U-Net抽出条件', 'Title または Abstract に "U-Net", "UNet", "nnU-Net", "U-Net++", "UNet++" を含む'])

    ws1.append([])
    headers = ['指標', '医療画像解析 (A0)', '製造欠陥検出 (B0)', '備考']
    ws1.append(headers)
    for i, h in enumerate(headers, 1):
        header_style(ws1.cell(ws1.max_row, i))

    a0s = summary['A0']
    b0s = summary['B0']
    rows_data = [
        ['全体Publication数（2015-2024）', a0s['total_valid'], b0s['total_valid'], ''],
        ['U-Net関連Publication数', a0s['unet_valid'], b0s['unet_valid'], ''],
        ['U-Net出現率（Publication）', f"{a0s['unet_valid']/a0s['total_valid']*100:.2f}%",
                                       f"{b0s['unet_valid']/b0s['total_valid']*100:.2f}%", ''],
        ['全体Family ID数', a0s['families'], b0s['families'], '同一発明の複数国出願を統合'],
        ['U-Net Family ID数', a0s['unet_families'], b0s['unet_families'], ''],
        ['U-Net出現率（Family ID）', f"{a0s['unet_families']/a0s['families']*100:.2f}%",
                                      f"{b0s['unet_families']/b0s['families']*100:.2f}%", ''],
        [],
        ['過去期間（2015-2018）', '', '', ''],
        ['　Publication数', a0s['past_total'], b0s['past_total'], ''],
        ['　U-Net Publication数', a0s['past_unet'], b0s['past_unet'], ''],
        ['　U-Net出現率', f"{a0s['past_unet']/max(a0s['past_total'],1)*100:.2f}%",
                          f"{b0s['past_unet']/max(b0s['past_total'],1)*100:.2f}%", ''],
        [],
        ['将来評価期間（2019-2024）', '', '', ''],
        ['　Publication数', a0s['future_total'], b0s['future_total'], ''],
        ['　U-Net Publication数', a0s['future_unet'], b0s['future_unet'], ''],
        ['　U-Net出現率', f"{a0s['future_unet']/max(a0s['future_total'],1)*100:.2f}%",
                           f"{b0s['future_unet']/max(b0s['future_total'],1)*100:.2f}%", ''],
    ]

    for row in rows_data:
        ws1.append(row)

    ws1.column_dimensions['A'].width = 38
    ws1.column_dimensions['B'].width = 22
    ws1.column_dimensions['C'].width = 22
    ws1.column_dimensions['D'].width = 30

    # ===== Sheet 2: 年次データ =====
    ws2 = wb.create_sheet('年次データ')
    years = list(range(2015, 2025))

    ws2['A1'] = '年次別U-Net出現データ'
    ws2['A1'].font = Font(bold=True, size=13, name='Calibri')

    ws2.append([])
    headers2 = ['年', 'A0 全体', 'A0 U-Net', 'A0 出現率(%)',
                 'B0 全体', 'B0 U-Net', 'B0 出現率(%)',
                 'Gap(A0-B0率)']
    ws2.append(headers2)
    for i, h in enumerate(headers2, 1):
        header_style(ws2.cell(ws2.max_row, i))

    a0_yr_t = a0s['yr_total']
    a0_yr_u = a0s['yr_unet']
    b0_yr_t = b0s['yr_total']
    b0_yr_u = b0s['yr_unet']

    period_labels = {2015: '過去', 2016: '過去', 2017: '過去', 2018: '過去',
                     2019: '将来', 2020: '将来', 2021: '将来', 2022: '将来', 2023: '将来', 2024: '将来'}

    for yr in years:
        a0t = a0_yr_t.get(str(yr), 0)
        a0u = a0_yr_u.get(str(yr), 0)
        b0t = b0_yr_t.get(str(yr), 0)
        b0u = b0_yr_u.get(str(yr), 0)
        a0r = a0u/a0t*100 if a0t > 0 else 0
        b0r = b0u/b0t*100 if b0t > 0 else 0
        gap = a0r - b0r

        row = ws2.max_row + 1
        ws2.cell(row, 1, yr)
        ws2.cell(row, 2, a0t)
        ws2.cell(row, 3, a0u)
        ws2.cell(row, 4, round(a0r, 2))
        ws2.cell(row, 5, b0t)
        ws2.cell(row, 6, b0u)
        ws2.cell(row, 7, round(b0r, 2))
        ws2.cell(row, 8, round(gap, 2))

        # Color code past vs future
        bg = 'EEF2F7' if yr <= 2018 else 'F0FFF0'
        for col in range(1, 9):
            c = ws2.cell(row, col)
            c.fill = PatternFill('solid', start_color=bg)
            c.font = Font(name='Calibri')
            c.alignment = Alignment(horizontal='center')

    for col in range(1, 9):
        ws2.column_dimensions[get_column_letter(col)].width = 15

    # ===== Sheet 3: 期間別比較 =====
    ws3 = wb.create_sheet('期間別比較（RQ2 RQ3）')

    ws3['A1'] = '期間別比較（RQ2: 過去期間, RQ3: 将来評価期間）'
    ws3['A1'].font = Font(bold=True, size=13, name='Calibri')

    ws3.append([])
    headers3 = ['期間', '分野', 'Publication数', 'U-Net Publication', '出現率(%)', 'Family ID数', 'U-Net Family', 'Family出現率(%)']
    ws3.append(headers3)
    for i, h in enumerate(headers3, 1):
        header_style(ws3.cell(ws3.max_row, i))

    period_data = [
        ('2015-2018（過去）', 'A0（医療画像解析）', a0s['past_total'], a0s['past_unet'],
         a0s['past_unet']/max(a0s['past_total'],1)*100, a0s['past_fam_total'], a0s['past_fam_unet'],
         a0s['past_fam_unet']/max(a0s['past_fam_total'],1)*100),
        ('2015-2018（過去）', 'B0（製造欠陥検出）', b0s['past_total'], b0s['past_unet'],
         b0s['past_unet']/max(b0s['past_total'],1)*100, b0s['past_fam_total'], b0s['past_fam_unet'],
         b0s['past_fam_unet']/max(b0s['past_fam_total'],1)*100),
        ('2019-2024（将来）', 'A0（医療画像解析）', a0s['future_total'], a0s['future_unet'],
         a0s['future_unet']/max(a0s['future_total'],1)*100, a0s['future_fam_total'], a0s['future_fam_unet'],
         a0s['future_fam_unet']/max(a0s['future_fam_total'],1)*100),
        ('2019-2024（将来）', 'B0（製造欠陥検出）', b0s['future_total'], b0s['future_unet'],
         b0s['future_unet']/max(b0s['future_total'],1)*100, b0s['future_fam_total'], b0s['future_fam_unet'],
         b0s['future_fam_unet']/max(b0s['future_fam_total'],1)*100),
    ]

    for prow in period_data:
        r = ws3.max_row + 1
        for i, v in enumerate(prow, 1):
            c = ws3.cell(r, i)
            if isinstance(v, float):
                c.value = round(v, 2)
                c.number_format = '0.00'
            else:
                c.value = v
            c.font = Font(name='Calibri')
            c.alignment = Alignment(horizontal='center')
            bg = 'EEF2F7' if '過去' in str(prow[0]) else 'F0FFF0'
            c.fill = PatternFill('solid', start_color=bg)

    for col in range(1, 9):
        ws3.column_dimensions[get_column_letter(col)].width = 20

    # ===== Sheet 4: 論文との照合 =====
    ws4 = wb.create_sheet('論文との照合')

    ws4['A1'] = '論文記載値との比較'
    ws4['A1'].font = Font(bold=True, size=13, name='Calibri')

    ws4.append([])
    headers4 = ['指標', '本抽出値', '論文記載値', '差異・備考']
    ws4.append(headers4)
    for i, h in enumerate(headers4, 1):
        header_style(ws4.cell(ws4.max_row, i))

    comp_rows = [
        ['A0 全体Publication', a0s['total_valid'], 15167, 'Combinedファイルには複数国出願が含まれるため約3倍'],
        ['A0 U-Net Publication', a0s['unet_valid'], 804, f'出現率: {a0s["unet_valid"]/a0s["total_valid"]*100:.2f}% (論文: 5.30%)'],
        ['A0 Family ID数', a0s['families'], 8513, ''],
        ['A0 U-Net Family', a0s['unet_families'], 577, f'Family出現率: {a0s["unet_families"]/a0s["families"]*100:.2f}% (論文: 6.78%)'],
        [],
        ['B0 全体Publication', b0s['total_valid'], 12212, ''],
        ['B0 U-Net Publication', b0s['unet_valid'], 72, f'出現率: {b0s["unet_valid"]/b0s["total_valid"]*100:.2f}% (論文: 0.59%)'],
        ['B0 Family ID数', b0s['families'], 8898, ''],
        ['B0 U-Net Family', b0s['unet_families'], 49, f'Family出現率: {b0s["unet_families"]/b0s["families"]*100:.2f}% (論文: 0.55%)'],
        [],
        ['A0 2015-2018 Publication', a0s['past_total'], 2515, ''],
        ['A0 2015-2018 U-Net', a0s['past_unet'], 25, f'出現率: {a0s["past_unet"]/max(a0s["past_total"],1)*100:.2f}% (論文: 0.99%)'],
        ['B0 2015-2018 Publication', b0s['past_total'], 2071, ''],
        ['B0 2015-2018 U-Net', b0s['past_unet'], 0, '論文と同様に過去期間はほぼゼロ'],
        [],
        ['A0 2019-2024 Publication', a0s['future_total'], 12652, ''],
        ['A0 2019-2024 U-Net', a0s['future_unet'], 779, f'出現率: {a0s["future_unet"]/max(a0s["future_total"],1)*100:.2f}% (論文: 6.16%)'],
        ['B0 2019-2024 Publication', b0s['future_total'], 10141, ''],
        ['B0 2019-2024 U-Net', b0s['future_unet'], 72, f'出現率: {b0s["future_unet"]/max(b0s["future_total"],1)*100:.2f}% (論文: 0.71%)'],
    ]

    for crow in comp_rows:
        if not crow:
            ws4.append([])
            continue
        r = ws4.max_row + 1
        for i, v in enumerate(crow, 1):
            c = ws4.cell(r, i, v)
            c.font = Font(name='Calibri')
            c.alignment = Alignment(horizontal='center' if i > 1 else 'left')

    ws4.column_dimensions['A'].width = 28
    ws4.column_dimensions['B'].width = 16
    ws4.column_dimensions['C'].width = 16
    ws4.column_dimensions['D'].width = 50

    wb.save(out_path)
    print(f'Excel saved: {out_path}')

if __name__ == '__main__':
    with open(WORKDIR / 'experiment_summary.json') as f:
        summary = json.load(f)

    a0_records = load_csv(WORKDIR / 'A0_data.csv')
    b0_records = load_csv(WORKDIR / 'B0_data.csv')

    out_path = WORKDIR / 'experiment_results.xlsx'
    make_excel(a0_records, b0_records, summary, out_path)
